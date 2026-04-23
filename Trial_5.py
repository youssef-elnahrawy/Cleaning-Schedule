# Imports
import streamlit as st
from icalendar import Calendar #used to parse .ics file
from openpyxl import load_workbook #used to read excel file
from datetime import date, timedelta #used to get current date
from openpyxl.styles import PatternFill, Fill, Color #used to color cells
from io import BytesIO

# Set up streamlit app
st.set_page_config(page_title="Airbnb Cleaning Schedule")
st.title("Airbnb Cleaning Schedule")
st.write("This app will help you create a cleaning schedule for your Airbnb property using a .ics file.")

uploaded_file = st.file_uploader("Upload your .ics file:", type=["ics"])

# Rest of normal code
START_ROW = 41
START_COL = 8
END_ROW = 4
GRID_WRAP_COL = 8
GRID_WRAP_ROW_OFFSET = 4
CALENDAR_START_ROW = 37
CALENDAR_END_ROW = 57
MAX_COLS = 9

SHEET_NAME_ROW = 1
SHEET_NAME_COL = 1

guest_fill = PatternFill(start_color="00CCFFCC", end_color="00CCFFCC", fill_type="solid")
cleaning_fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type="solid")
white_fill = PatternFill(start_color="00FFFFFF", end_color="00FFFFFF", fill_type="solid")



# Load excel template for calendar
wb = load_workbook(filename='template.xlsx')
sheet = wb.active

# Read the .ics file
if uploaded_file is not None:
    ics_content = uploaded_file.read()

    # Parse the .ics file
    calendar = Calendar.from_ical(ics_content)

    # Add the days to the calendar
    current_day_name = date.today().weekday() + 1
    number_of_days = 70 - current_day_name
    row = START_ROW
    col = START_COL
    while row >= END_ROW:
        cell = sheet.cell(row=row, column=col)
        cell.value = date.today() + timedelta(days=number_of_days)
        number_of_days -= 1

        for event in calendar.walk('vevent'):
            event_end = event.get('dtend')
            event_start = event.get('dtstart')
            if event_end and event_start and cell.value == event_end.dt and event.get('summary') == "Reserved":
                cell2 = sheet.cell(row=row + 1, column=col)
                cell2.value = "Sale Huesped - 1pm"

                start_date = event_start
                end_date = event_end
                length = (end_date.dt - start_date.dt).days

                y = 0
                col2 = col
                row2 = row + 1
                while length + 1 > 0:
                    cell2 = sheet.cell(row=row2, column=col2 + y)
                    cell2.fill = guest_fill
                    y -= 1
                    length -= 1
                    if col2 + y == 1:
                        col2 += 7
                        row2 -= GRID_WRAP_ROW_OFFSET
                    if row2 <= 5:
                        break

                col3 = col + 1
                row3 = row + 2
                if col3 > GRID_WRAP_COL:
                    col3 -= 7
                    row3 += GRID_WRAP_ROW_OFFSET

                cell3 = sheet.cell(row=row3, column=col3)
                cell3.value = "LIMPIEZA"
                cell3.fill = cleaning_fill

            if event_end and event_start and cell.value == event_start.dt and event.get('summary') == "Reserved":
                cell3 = sheet.cell(row=row + 1, column=col)

                if cell3.value == "Sale Huesped - 1pm":
                    cell3.value = "Sale/Entra Huesped"
                else:
                    cell3.value = "Entra Huesped - 4pm"

                row3 = row + 2
                col3 = col + 1
                cell3 = sheet.cell(row=row + 1, column=col)

                if col3 > GRID_WRAP_COL:
                    col3 -= 7
                    row3 += GRID_WRAP_ROW_OFFSET

                cell4 = sheet.cell(row=row3, column=col3)

                if cell3.value == "Sale/Entra Huesped":
                    cell4 = sheet.cell(row=row + 2, column=col)
                    cell4.value = "LIMPIEZA (1pm - 4pm)"
                    cell4.fill = cleaning_fill

                    col4 = col + 1
                    row4 = row + 2
                    if col4 > GRID_WRAP_COL:
                        col4 -= 7
                        row4 += GRID_WRAP_ROW_OFFSET
                    cell4 = sheet.cell(row=row4, column=col4)
                    cell4.value = ""
                    cell4.fill = white_fill

        col -= 1
        if col == 1:
            row -= GRID_WRAP_ROW_OFFSET
            col = GRID_WRAP_COL

    for col in range(1, MAX_COLS):
        for row in range(CALENDAR_START_ROW, CALENDAR_END_ROW):
            cell = sheet.cell(row=row, column=col)
            cell.fill = white_fill
            cell.value = ""
    sheet.cell(row=SHEET_NAME_ROW, column=SHEET_NAME_COL).value = ""

else:
    st.write(" ")
    st.error("Please upload a valid .ics file.")


output = BytesIO()
wb.save(output)

st.download_button(
    label="Download Cleaning Schedule",
    data=output.getvalue(),
    file_name="calendario_de_limpieza.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)