from openpyxl import Workbook
from openpyxl.styles import PatternFill, Fill, Color

workbook = Workbook()
sheet = workbook.active

for x in range(10):
    sheet.merge_cells("D2:F2")

sheet["D2"] = "/ CALENDARIO DE LIMIEZA"

sheet["B4"] = "Domingo"
sheet["C4"] = "Lunes"
sheet["D4"] = "Martes"
sheet["E4"] = "Miercoles"
sheet["F4"] = "Jueves"
sheet["G4"] = "Viernes"
sheet["H4"] = "Sabado"
sheet["AZ100"] = 0

fill1 = PatternFill(start_color="0099CCFF", end_color="0099CCFF", fill_type="solid")
# Apply the fill to the range B4:H4
for col in range(2, 9):  # Columns B to H are 2 to 8 in 1-indexed
    cell = sheet.cell(row=4, column=col)
    cell.fill = fill1

fill2 = PatternFill(start_color="00FFFFFF", end_color="00FFFFFF", fill_type="solid")

for row in range(1, sheet.max_row + 1):
    cell = sheet.cell(row=row, column=1)
    cell.fill = fill2

for col in range(1, sheet.max_column + 1):
    for row in range(1, 4):
        cell = sheet.cell(row=row, column=col)
        cell.fill = fill2

for col in range(9, sheet.max_column + 1):
    for row in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=col)
        cell.fill = fill2

for col in range(1, sheet.max_column + 1):
    for row in range(37, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=col)
        cell.fill = fill2

fill3 = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")

for col in range(2, 9):
    for row in [5, 9, 13, 17, 21, 25, 29, 33]:
        cell = sheet.cell(row=row, column=col)
        cell.fill = fill3




workbook.save(filename="Trial 1 File.xlsx")