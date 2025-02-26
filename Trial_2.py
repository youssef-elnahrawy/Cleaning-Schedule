from icalendar import Calendar #used to parse .ics file
import tkinter as tk #used for prompting user to uplad file
from tkinter import filedialog #used for prompting user to uplad file
from openpyxl import load_workbook #used to read excel file
from datetime import date, timedelta #used to get current date
from openpyxl.styles import PatternFill, Fill, Color #used to color cells

fill1 = PatternFill(start_color="00CCFFCC", end_color="00CCFFCC", fill_type="solid") #green
fill2 = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type="solid") #yellow
fill3 = PatternFill(start_color="00FFFFFF", end_color="00FFFFFF", fill_type="solid") #white

# Prompt user to upload .ics file
root = tk.Tk()
root.withdraw()
ics_file_path = filedialog.askopenfilename()

# Load excel template for calendar
wb = load_workbook(filename = 'template.xlsx')
sheet = wb.active

# Read the .ics file
with open(ics_file_path, 'rb') as f:
    ics_content = f.read()

# Parse the .ics file
calendar = Calendar.from_ical(ics_content)

# Add the days to the calendar
number_of_days = 0
for row in [5, 9, 13, 17, 21, 25, 29, 33]:
    for col in range(2, 9):
        cell = sheet.cell(row=row, column=col)
        cell.value = date.today() + timedelta(days=number_of_days)
        number_of_days += 1

# Add events to calendar/highlight cells
        for event in calendar.walk('vevent'):
            if cell.value == event.get('dtstart').dt and event.get('summary') == "Reserved":
                
                # summary = event.get('summary')
                start_date = event.get('dtstart')
                end_date = event.get('dtend')
                length = (end_date.dt - start_date.dt).days
                
                # print("Event: ", summary)
                # print("Start Date: ", start_date.dt)
                # print("End Date: ", end_date.dt)
                # print("Length: ", length)
                # print()

                # add beginning of event name to the cell
                cell2 = sheet.cell(row=row+1, column=col)
                cell3 = sheet.cell(row=row+2, column=col+1)
                
                # If end of event overlaps start of another event
                if cell2.value == "Sale Huesped - 1pm":
                    cell2.value = "Sale/Entra Huesped"
                    cell3 = sheet.cell(row=row+2, column=col)
                    cell3.value = "LIMPIEZA (1pm - 4pm)"
                    cell3.fill = fill2
                    cell3 = sheet.cell(row=row+2, column=col+1)
                    if col + 1 == 9:
                        col -= 7
                        row += 4
                        cell3.value = " "
                        cell3.fill = fill3
                        col += 7
                        row -= 4
                    cell3.value = ""
                    cell3.fill = fill3
                else:
                    cell2.value = "Entra Huesped - 4pm"

                # color the relevent cells for the event
                y = 0
                row3 = row + 1
                col3 = col
                while length + 1 > 0:
                    cell2 = sheet.cell(row = row3, column = col3 + y)
                    cell2.fill = fill1
                    if col3 + y == 8:
                        col3 -= 7
                        row3 += 4
                    
                    if length == 0:
                        cell2.value = "Sale Huesped - 1pm" #add event name to the last cell of the event

                        cell4 = sheet.cell(row = row3 + 1, column = col3 + y)
                        if cell4.value != "LIMPIEZA (1pm - 4pm)":
                            cell3 = sheet.cell(row = row3 + 1, column = col3 + y + 1)
                            cell3.value = "LIMPIEZA"
                            cell3.fill = fill2

                    y += 1
                    length -= 1
                

                
# Add cleaning schedule to calendar
            # if cell.value == event.get('dtend').dt:
            #     fill2 = PatternFill(start_color="00FFCC00", end_color="00FFCC00", fill_type="solid")
            #     cell3 = sheet.cell(row=row+2, column=col+1)
            #     if col + 1 == 8:
            #             col -= 7
            #             row += 4
            #     cell3.value = "LIMPIEZA"
            #     cell3.fill = fill2




# # Iterate through the events in the calendar
# for event in calendar.walk('vevent'):

#     start_date = event.get('dtstart')
#     end_date = event.get('dtend')
#     summary = event.get('summary')

#     print ("Event: ", summary)
#     print ("Start Date: ", start_date.dt)
#     print ("End Date: ", end_date.dt)
#     print()

wb.save(filename="calendario_de_limieza.xlsx")
